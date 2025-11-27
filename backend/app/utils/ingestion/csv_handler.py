import csv
from pathlib import Path
from typing import Generator

from openpyxl import load_workbook

from .base import IngestionHandler, StreamedDocument


class TabularHandler(IngestionHandler):
    supported_types = [".csv", ".tsv", ".xlsx", ".xls"]

    def stream(self, source: Path) -> StreamedDocument:
        suffix = source.suffix.lower()
        if suffix in (".xlsx", ".xls"):
            iterator = self._iter_excel(source)
        else:
            delimiter = "\t" if suffix == ".tsv" else ","
            iterator = self._iter_csv(source, delimiter)
        return StreamedDocument(chunks=iterator, metadata={"source": str(source), "type": "tabular"})

    def _iter_csv(self, source: Path, delimiter: str) -> Generator[str, None, None]:
        with source.open("r", newline="", encoding="utf-8", errors="ignore") as handle:
            reader = csv.reader(handle, delimiter=delimiter)
            for row in reader:
                yield delimiter.join(row)

    def _iter_excel(self, source: Path) -> Generator[str, None, None]:
        workbook = load_workbook(filename=source, read_only=True, data_only=True)
        try:
            for sheet in workbook.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    values = ["" if cell is None else str(cell) for cell in row]
                    yield ",".join(values)
        finally:
            workbook.close()


